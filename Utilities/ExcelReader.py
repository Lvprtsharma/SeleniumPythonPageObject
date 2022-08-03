import openpyxl


def getrowcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def getcolcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column


def getcelldata(path, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=colnum).value


def setcelldata(path, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(path)


patH = "excel\\testdata.xlsx"
sheetName = "Passwords"

rows = getrowcount(patH, sheetName)
cols = getcolcount(patH, sheetName)

print(getcelldata(patH, sheetName, 2, 3))

setcelldata(patH, sheetName, 5, 1, "DOB")

print(rows, "---", cols)
